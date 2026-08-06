"""
Reorganiza el output del run_sensibilidades en un xlsx multi-hoja organizado:
- Resumen con baseline + Top 5 Up/Down + Ranking por bloque
- Tabla detallada (puede dividirse por grupo si tienes muchos outputs)
- Series mensuales por grupo

USO:
1. Editar SRC con la ruta del xlsm de salida del script PowerShell
2. Editar DST con la ruta donde guardar el xlsx organizado
3. Ajustar CATEGORIES y GROUPS según los escenarios y outputs de tu modelo
4. python build_organized.py

REQUIERE: openpyxl
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# =============================================================================
# CONFIG - AJUSTAR
# =============================================================================
HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "MODELO_RESULTADOS_SENSIBILIDADES.xlsm")  # output del run_sensibilidades.ps1
DST = os.path.join(HERE, "RESULTADOS_ORGANIZADO.xlsx")
SHEET_RESULTADOS = "Reusltado Sensibilidades"  # nombre exacto en el xlsm fuente

# Categorías de escenarios (por idx — empieza en 1, baseline=0)
# Si solo tienes 16 escenarios estándar, ajustá a esos índices
CATEGORIES = [
    {"name": "A. Fisher Taylor (IPC + IBR + IPP factor coupled)", "scenarios": [1,2,3,4,5,6]},
    {"name": "B. IPC puro (sin coupling)",                        "scenarios": [7,8,9,10,11,12]},
    {"name": "C. IBR puro",                                        "scenarios": [13,14,15,16,17,18]},
    {"name": "D. Factor IPP-IPC (spread)",                         "scenarios": [19,20,21,22,23,24,25,26]},
    {"name": "E. Ke / Rf",                                         "scenarios": [27,28,29,30,31,32,33,34]},
    {"name": "F. CAPM componentes",                                "scenarios": [35,36,37,38,39,40,41,42]},
    {"name": "G. OPEX",                                            "scenarios": [43,44,45,46,47,48]},
    {"name": "J. Mantenimiento",                                   "scenarios": [49,50]},
    {"name": "K. Combinados",                                       "scenarios": [51,52,53,54,55]}
]

# Grupos de outputs (para dividir tabla detallada en múltiples hojas si son muchos)
# Si tenés pocos outputs, dejá un solo grupo con todos.
# El script detecta automáticamente los outputs disponibles en el header de la fila 12.
GROUPS = [
    ("Valoración + Caja", ["Equity DDM", "IRR Eq", "CFADS Total", "FCFE Total", "PR Total", "Caja Final", "Caja Min"]),
    # Agregar más grupos si tenés muchos outputs:
    # ("Ingresos", ["Ingresos Total", "Ing Ecopetrol", "Ing ODL", ...]),
    # ("Costos", ["Costos Op Total", "Cost Personal", ...]),
]

# Series mensuales por grupo (busca por label en bloques [X MENSUAL] del xlsm)
SERIES_GROUPS = [
    ("Cash Flow", ["CFADS", "FCFE (Disp Acc)", "Pagos Restringidos", "Saldo Caja"])
]

# =============================================================================
# STYLING
# =============================================================================
NAVY = "1F3864"; TEAL = "2E8B8B"; GREEN_POS = "C6EFCE"; RED_NEG = "FFC7CE"; WHITE = "FFFFFF"
THIN = Side(border_style="thin", color="999999")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr_style(cell, bg=NAVY, color=WHITE):
    cell.font = Font(name="Arial", size=10, bold=True, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_ALL

def title_style(cell, size=14, color=NAVY):
    cell.font = Font(name="Arial", size=size, bold=True, color=color)

# =============================================================================
# READ SOURCE
# =============================================================================
print(f"Leyendo: {SRC}")
wb_src = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_src = wb_src[SHEET_RESULTADOS]

# Auto-detect outputs from header row 12 (col 4 onwards, before "Audit")
out_to_col = {}
for c in range(4, 200):
    h = ws_src.cell(row=12, column=c).value
    if h and isinstance(h, str) and not h.startswith("Δ") and not h.startswith("Delta") and h != "Audit":
        out_to_col[h] = c
print(f"Outputs detectados ({len(out_to_col)}): {list(out_to_col.keys())}")

# Read summary rows
audit_col = max(out_to_col.values()) + 2 if out_to_col else 4
summary = []
for r in range(13, 100):
    idx = ws_src.cell(row=r, column=1).value
    if idx is None: continue
    row = {"idx": idx, "name": ws_src.cell(row=r, column=2).value, "notes": ws_src.cell(row=r, column=3).value}
    audit_v = None
    for ac in range(audit_col, audit_col + 5):
        v = ws_src.cell(row=r, column=ac).value
        if v and isinstance(v, str): audit_v = v; break
    row["audit"] = audit_v
    for label, col in out_to_col.items():
        row[label] = ws_src.cell(row=r, column=col).value
        row[label + "_pct"] = ws_src.cell(row=r, column=col+1).value
    summary.append(row)
print(f"Escenarios: {len(summary)}")
baseline = summary[0] if summary else None

# Read series blocks
series_data = {}
for r in range(70, ws_src.max_row + 1):
    v = ws_src.cell(row=r, column=1).value
    if v and isinstance(v, str) and "MENSUAL" in v.upper() and "[" in v:
        title = v.replace("[", "").replace("]", "").replace(" MENSUAL", "").strip()
        time_row = r + 1
        data_start = r + 2
        time_labels = []
        for c in range(3, 350):
            tv = ws_src.cell(row=time_row, column=c).value
            if tv is None and c > 50: break
            time_labels.append(tv)
        rows = []
        for i in range(len(summary)):
            scen_data = []
            scen_name = ws_src.cell(row=data_start + i, column=2).value
            for c in range(3, 3 + len(time_labels)):
                scen_data.append(ws_src.cell(row=data_start + i, column=c).value)
            rows.append({"name": scen_name, "data": scen_data})
        series_data[title] = {"time": time_labels, "rows": rows}
print(f"Series blocks: {len(series_data)}")

# =============================================================================
# BUILD WORKBOOK
# =============================================================================
print("\nConstruyendo...")
wb_new = Workbook()
wb_new.remove(wb_new.active)

# SHEET 1: Resumen
ws = wb_new.create_sheet("1. Resumen")
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 38
for c in ['C','D','E','F']: ws.column_dimensions[c].width = 14

ws['B2'] = "Análisis de Sensibilidad"
title_style(ws['B2'], size=14)
if baseline:
    primary = list(out_to_col.keys())[0]
    ws['B3'] = f"{len(summary)-1} escenarios · BASELINE {primary} = {baseline.get(primary, 0):,.0f} · {len(out_to_col)} outputs"
    ws['B3'].font = Font(name="Arial", size=10, italic=True, color="8497B0")

# Baseline metrics (todos los outputs)
ws['B5'] = "Métricas Baseline"; title_style(ws['B5'], size=11)
for i, (label, _) in enumerate(out_to_col.items()):
    r = 6 + i
    ws.cell(row=r, column=2, value=label).font = Font(name="Arial", size=10, bold=True)
    val = baseline.get(label) if baseline else None
    if val is not None:
        c = ws.cell(row=r, column=3, value=val)
        c.number_format = "0.00%" if "IRR" in label else "#,##0"

# Top movers (por primer output)
if baseline and len(summary) > 1:
    primary = list(out_to_col.keys())[0]
    base_p = baseline.get(primary, 0)
    if base_p:
        scen_d = []
        for row in summary[1:]:
            v = row.get(primary); pct = row.get(primary + "_pct")
            if v is not None and pct is not None:
                scen_d.append((row['idx'], row['name'], v, v - base_p, pct))

        # Top Up
        start_up = 8 + len(out_to_col)
        ws.cell(row=start_up, column=2, value=f"Top 5 Upside ({primary})")
        title_style(ws.cell(row=start_up, column=2), size=11, color="2E7D32")
        for i, h in enumerate(["#", "Escenario", primary, "Δ abs", "Δ %"]):
            hdr_style(ws.cell(row=start_up+1, column=2+i, value=h), bg="2E7D32")
        for i, (idx, name, v, dabs, dpct) in enumerate(sorted(scen_d, key=lambda x: x[4], reverse=True)[:5]):
            r = start_up + 2 + i
            ws.cell(row=r, column=2, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=3, value=name)
            c = ws.cell(row=r, column=4, value=v); c.number_format = "#,##0"
            c = ws.cell(row=r, column=5, value=dabs); c.number_format = "+#,##0;-#,##0"; c.font = Font(name="Arial", size=10, bold=True, color="2E7D32")
            c = ws.cell(row=r, column=6, value=dpct); c.number_format = "+0.00%;-0.00%"; c.font = Font(name="Arial", size=10, bold=True, color="2E7D32")
            for col in range(2, 7): ws.cell(row=r, column=col).border = BORDER_ALL

        # Top Down
        start_dn = start_up + 9
        ws.cell(row=start_dn, column=2, value=f"Top 5 Downside ({primary})")
        title_style(ws.cell(row=start_dn, column=2), size=11, color="C62828")
        for i, h in enumerate(["#", "Escenario", primary, "Δ abs", "Δ %"]):
            hdr_style(ws.cell(row=start_dn+1, column=2+i, value=h), bg="C62828")
        for i, (idx, name, v, dabs, dpct) in enumerate(sorted(scen_d, key=lambda x: x[4])[:5]):
            r = start_dn + 2 + i
            ws.cell(row=r, column=2, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=3, value=name)
            c = ws.cell(row=r, column=4, value=v); c.number_format = "#,##0"
            c = ws.cell(row=r, column=5, value=dabs); c.number_format = "+#,##0;-#,##0"; c.font = Font(name="Arial", size=10, bold=True, color="C62828")
            c = ws.cell(row=r, column=6, value=dpct); c.number_format = "+0.00%;-0.00%"; c.font = Font(name="Arial", size=10, bold=True, color="C62828")
            for col in range(2, 7): ws.cell(row=r, column=col).border = BORDER_ALL

# SHEETS 2..N: Tablas detalladas por grupo
for grp_idx, (grp_name, grp_outputs) in enumerate(GROUPS):
    sheet_name = f"{grp_idx+2}. {grp_name}"[:31]
    ws = wb_new.create_sheet(sheet_name)
    grp_outputs_existing = [o for o in grp_outputs if o in out_to_col]
    if not grp_outputs_existing:
        ws['A1'] = f"(No hay outputs del grupo {grp_name} en el archivo)"
        continue

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 38
    audit_col_local = 4 + len(grp_outputs_existing)*2
    for c in range(4, audit_col_local + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.cell(row=1, column=1, value=f"{grp_name} — {len(summary)} escenarios")
    title_style(ws['A1'], size=12)

    hdr_style(ws.cell(row=3, column=1, value="#"))
    hdr_style(ws.cell(row=3, column=2, value="Escenario"))
    hdr_style(ws.cell(row=3, column=3, value="Notas"))
    col = 4
    for out in grp_outputs_existing:
        hdr_style(ws.cell(row=3, column=col, value=out))
        hdr_style(ws.cell(row=3, column=col+1, value="Δ%"))
        col += 2
    hdr_style(ws.cell(row=3, column=audit_col_local, value="Audit"))

    # BASELINE
    if baseline:
        ws.cell(row=4, column=1, value=0)
        ws.cell(row=4, column=2, value="BASELINE")
        ws.cell(row=4, column=3, value="Sin shocks")
        col = 4
        for out in grp_outputs_existing:
            v = baseline.get(out)
            c = ws.cell(row=4, column=col, value=v); c.number_format = "0.00%" if "IRR" in out else "#,##0"
            c2 = ws.cell(row=4, column=col+1, value=0); c2.number_format = "0.00%"
            col += 2
        ws.cell(row=4, column=audit_col_local, value="BASE")
        for c_idx in range(1, audit_col_local + 1):
            cc = ws.cell(row=4, column=c_idx)
            cc.fill = PatternFill("solid", fgColor="DDEBF7")
            cc.font = Font(name="Arial", size=9, bold=True)
            cc.border = BORDER_ALL

    # Bloques con headers de categoria
    current_row = 6
    for cat in CATEGORIES:
        c = ws.cell(row=current_row, column=1, value=cat['name'])
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=audit_col_local)
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        for sidx in cat["scenarios"]:
            if sidx <= len(summary) - 1:
                row_data = summary[sidx]
                ws.cell(row=current_row, column=1, value=row_data['idx']).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=2, value=row_data['name'])
                ws.cell(row=current_row, column=3, value=row_data['notes'])
                col = 4
                for out in grp_outputs_existing:
                    v = row_data.get(out); v_pct = row_data.get(out + "_pct")
                    c1 = ws.cell(row=current_row, column=col, value=v)
                    c1.number_format = "0.00%" if "IRR" in out else "#,##0"
                    c1.font = Font(name="Arial", size=9); c1.border = BORDER_ALL
                    c2 = ws.cell(row=current_row, column=col+1, value=v_pct)
                    c2.number_format = "0.00%"; c2.border = BORDER_ALL
                    if v_pct is not None:
                        if v_pct > 0.001:
                            c2.fill = PatternFill("solid", fgColor=GREEN_POS)
                            c2.font = Font(name="Arial", size=9, bold=True, color="2E7D32")
                        elif v_pct < -0.001:
                            c2.fill = PatternFill("solid", fgColor=RED_NEG)
                            c2.font = Font(name="Arial", size=9, bold=True, color="C62828")
                    col += 2
                audit = row_data.get('audit') or "(falta)"
                c = ws.cell(row=current_row, column=audit_col_local, value=audit)
                c.font = Font(name="Arial", size=8); c.border = BORDER_ALL
                if isinstance(audit, str) and "WARN" in audit:
                    c.fill = PatternFill("solid", fgColor="FFE699")
                elif isinstance(audit, str) and ("ERROR" in audit or "falta" in audit):
                    c.fill = PatternFill("solid", fgColor="FFC7CE")
                for c_idx in [1, 2, 3]:
                    cc = ws.cell(row=current_row, column=c_idx)
                    cc.font = Font(name="Arial", size=9); cc.border = BORDER_ALL
                    cc.alignment = Alignment(horizontal="center" if c_idx == 1 else "left")
                current_row += 1
        current_row += 1
    ws.freeze_panes = "D4"

# SHEETS series mensuales por grupo
for grp_idx, (grp_name, grp_series) in enumerate(SERIES_GROUPS):
    sheet_name = f"{len(GROUPS)+grp_idx+2}. Series {grp_name}"[:31]
    ws = wb_new.create_sheet(sheet_name)
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 36
    for c in range(3, 350): ws.column_dimensions[get_column_letter(c)].width = 11

    ws.cell(row=1, column=1, value=f"Series Mensuales — {grp_name}")
    title_style(ws['A1'], size=12)

    current_row = 3
    for series_label in grp_series:
        if series_label not in series_data:
            continue
        block = series_data[series_label]
        c = ws.cell(row=current_row, column=1, value=f"[{series_label}]")
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        hdr_style(ws.cell(row=current_row, column=1, value="#"))
        hdr_style(ws.cell(row=current_row, column=2, value="Escenario"))
        for j, t in enumerate(block['time']):
            cc = ws.cell(row=current_row, column=3+j, value=t)
            cc.font = Font(name="Arial", size=8, bold=True, color=WHITE)
            cc.fill = PatternFill("solid", fgColor=NAVY)
            cc.alignment = Alignment(horizontal="center"); cc.border = BORDER_ALL
            if isinstance(t, (int, float)) and t > 1900: cc.number_format = "0"
        current_row += 1

        for i, scenrow in enumerate(block['rows']):
            ws.cell(row=current_row, column=1, value=i)
            ws.cell(row=current_row, column=2, value=scenrow['name'])
            is_baseline = (i == 0)
            for j, v in enumerate(scenrow['data']):
                cc = ws.cell(row=current_row, column=3+j, value=v)
                cc.font = Font(name="Arial", size=8, bold=is_baseline)
                cc.number_format = "#,##0;-#,##0;-"
                cc.alignment = Alignment(horizontal="right"); cc.border = BORDER_ALL
                if is_baseline: cc.fill = PatternFill("solid", fgColor="DDEBF7")
            for col in [1, 2]:
                cc = ws.cell(row=current_row, column=col)
                cc.font = Font(name="Arial", size=8, bold=is_baseline)
                cc.alignment = Alignment(horizontal="center" if col == 1 else "left"); cc.border = BORDER_ALL
                if is_baseline: cc.fill = PatternFill("solid", fgColor="DDEBF7")
            current_row += 1
        current_row += 2
    ws.freeze_panes = "C3"

print(f"\nGuardando {DST}")
wb_new.save(DST)
print("DONE")
